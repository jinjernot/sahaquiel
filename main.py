import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


def load_raw():
    try:
        df = pd.read_excel("RAW.xlsm")
        return df
    except FileNotFoundError as e:
        print("Error: File not found:", e)
        return None

def main():
    df = load_raw()
    #df = df.iloc[:, :47]  # Delete columns from index 47 to the end
    if df is not None:
        workbook = load_workbook("RAW.xlsm")
        writer = pd.ExcelWriter("new_file.xlsm", engine="openpyxl")

        # Write DataFrame to the sheet without the index column and header
        df.to_excel(writer, sheet_name="Instructions", index=False, header=False)

        # Access the workbook and sheet
        sheet = writer.book["Instructions"]

        # Insert three rows at the start of the sheet
        sheet.insert_rows(1, amount=4)

        # Apply formatting to the inserted rows and first 4 rows
        fillH = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        font = Font(color="FFFFFF", size=26)
        
        for row in sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=48):
            for cell in row:
                cell.fill = fillH
                cell.font = font
                sheet.row_dimensions[cell.row].height = 30

        # Apply formatting to the range A11:J11 and adjust column widths
        fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        for row in sheet.iter_rows(min_row=11, max_row=11, min_col=1, max_col=10):
            smallfont = Font(color="FFFFFF", size=10)
            for cell in row:
                cell.fill = fill
                cell.font = smallfont
                cell.alignment = cell.alignment.copy(wrap_text=True)  #
                cell.column_letter
                sheet.column_dimensions[cell.column_letter].width = 30
                cell.row
                sheet.row_dimensions[cell.row].height = 60
        
        smallblackfont = Font(color="000000", size=10)
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        for row in sheet.iter_rows(min_row=11, max_row=11, min_col=11, max_col=48):
            for cell in row:
                cell.font = smallblackfont
                cell.font = cell.font.copy(bold=True)
                cell.alignment = center_alignment

        # Specify the row number to modify
        row_num = 7

        # Specify the desired height
        height = 60

        # Set the height for the specified row
        sheet.row_dimensions[row_num].height = height
        # Set the alignment for the specified row
        alignment = Alignment(horizontal='center', vertical='center')

        for cell in sheet[row_num]:
            cell.alignment = alignment

        # Define the range to merge
        merge_range = "K8:O8"

        # Merge the cells
        sheet.merge_cells(merge_range)

        # Set the alignment for the merged cell
        merged_cell = sheet["K8"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Define the range to merge
        merge_range = "P7:U7"

        # Merge the cells
        sheet.merge_cells(merge_range)

        # Set the alignment for the merged cell
        merged_cell = sheet["P7"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Define the range to merge
        merge_range = "X7:AB7"

        # Merge the cells
        sheet.merge_cells(merge_range)

        # Set the alignment for the merged cell
        merged_cell = sheet["X7"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Define the range to merge
        merge_range = "AF7:AK7"

        # Merge the cells
        sheet.merge_cells(merge_range)

        # Set the alignment for the merged cell
        merged_cell = sheet["AF7"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')


        # Define the range to merge
        merge_range = "P5:W6"

        # Get the value of cell S6
        cell_S6_value = sheet["S6"].value

        # Merge the cells
        sheet.merge_cells(merge_range)

        # Set the alignment for the merged cell
        merged_cell = sheet["P5"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')


        # Assign the value of cell S6 to the merged cell
        merged_cell.value = cell_S6_value

        # Define the range to merge
        merge_range = "X5:AE6"

        # Get the value of cell AA6
        cell_AA6_value = sheet["AA6"].value

        # Merge the cells
        sheet.merge_cells(merge_range)

        # Set the alignment for the merged cell
        merged_cell = sheet["X5"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Assign the value of cell AA6 to the merged cell
        merged_cell.value = cell_AA6_value


        # Define the range to merge
        merge_range = "AF5:AM6"

        # Get the value of cell AI6
        cell_AI6_value = sheet["AI6"].value

        # Merge the cells
        sheet.merge_cells(merge_range)

        # Set the alignment for the merged cell
        merged_cell = sheet["AF5"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Assign the value of cell AI6 to the merged cell
        merged_cell.value = cell_AI6_value

        # Define the range to merge
        merge_range = "AN5:AV6"

        # Get the value of cell AQ6
        cell_AQ6_value = sheet["AQ6"].value

        # Merge the cells
        sheet.merge_cells(merge_range)

        # Set the alignment for the merged cell
        merged_cell = sheet["AN5"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Assign the value of cell AQ6 to the merged cell
        merged_cell.value = cell_AQ6_value

        merge_range = "AN7:AS7"

        # Merge the cells
        sheet.merge_cells(merge_range)

        # Set the alignment for the merged cell
        merged_cell = sheet["AN7"]
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')



        # Select the ranges
        Box_Info = sheet["K8:O11"]
        Standard_AIR_IP = sheet["P5:W11"]
        Standard_TRUCK_IP = sheet["X5:AE11"]
        Standard_SEA_IP = sheet["AF5:AM11"]
        Customer_Specific_EP = sheet["AN5:AV11"]

        # Define the background colors for each range
        colors = {
            Box_Info: "e6e1eb",
            Standard_AIR_IP: "4e9fd5",
            Standard_TRUCK_IP: "a9cdf2",
            Standard_SEA_IP: "dbe6f6",
            Customer_Specific_EP: "34d1da",
        }

        # Define the border style
        border_style = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )

        # Apply background colors and borders to each range
        for range_, color in colors.items():
            for row in range_:
                for cell in row:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.border = border_style
               
        # Define border style
        #border_style = Side(border_style="thin", color="000000")

        #Apply border to rows starting from 9 and below
        #for row_num in range(12, sheet.max_row + 1):
         #   for col_num in range(1, sheet.max_column + 1):
         #       cell = sheet.cell(row=row_num, column=col_num)
         #       cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

        # Insert text into cell A3 and increase font size
        sheet["A3"] = "PSG EMEA PRODUCTS Weight, Dimensions, Pallet data"
        #sheet["A3"].font = Font(size=16)
        # Set background color
        
        # Insert text into cells
        fontred = Font(color="FF0000")
        sheet["A6"].font = fontred
        sheet["A6"] = "HP Confidential. For HP and Partner internal use only"
        sheet["A7"] = "The information contained herein is HP Confidential. Disclosure is governed by your HP Partner Agreement, HP Retail Partner Agreement, or another applicable contract (e.g., Confidential Disclosure Agreement)."
        
        # Delete content from cell A5
        sheet["A5"].value = None

        # Add autofilters to row 11
        sheet.auto_filter.ref = "A11:AV11"   

        # Define the border style
        border_style_top = Border(top=Side(border_style="thick"))
        border_style_bottom = Border(bottom=Side(border_style="thick"))
        border_style_left = Border(left=Side(border_style="thick"))

       

        fill = PatternFill(start_color="f6eddc", end_color="f6eddc", fill_type="solid")

        # Apply background color to rows 9 and 10
        for row_num in range(9, 11):
            for col_num in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.fill = fill

        # Save the workbook
        writer.save()
        writer.close()

if __name__ == "__main__":
    main()
